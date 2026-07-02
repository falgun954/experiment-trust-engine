import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"
raw = pd.read_csv(Path(__file__).resolve().parents[1] / "data/processed/excel_raw_inputs.csv")
metric_defs = pd.read_csv(Path(__file__).resolve().parents[1] / "data/raw/metric_definitions.csv")
daily_wide = pd.read_csv(Path(__file__).resolve().parents[1] / "data/processed/daily_lift_wide.csv")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT, color="0000FF", size=10)          # blue = hardcoded input
FORMULA_FONT = Font(name=FONT, color="000000", size=10)        # black = formula
LINK_FONT = Font(name=FONT, color="008000", size=10)           # green = cross-sheet link
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F2937")
SUBTITLE_FONT = Font(name=FONT, size=10, color="6B7280", italic=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)

# =====================================================================
# SHEET 1: Raw_Inputs — hardcoded values, blue font, sourced
# =====================================================================
ws = wb.create_sheet("Raw_Inputs")
cols = ["experiment_id", "experiment_name", "owner_team", "control_n", "treatment_n",
        "control_conversions", "treatment_conversions", "control_guardrail_n", "treatment_guardrail_n",
        "active_users_def_v1", "active_users_def_v2", "required_sample_size_per_arm",
        "novelty_decay_pct_per_day", "is_novelty_effect", "start_date", "end_date"]
headers = ["Experiment ID", "Experiment Name", "Owner Team", "Control N", "Treatment N",
           "Control Conversions", "Treatment Conversions", "Control Guardrail N", "Treatment Guardrail N",
           "Active Users (Def v1)", "Active Users (Def v2)", "Required N / Arm (MDE=2pp, power=80%)",
           "Novelty Decay %/day", "Is Novelty Effect (0/1)", "Start Date", "End Date"]

for j, h in enumerate(headers, 1):
    c = ws.cell(1, j, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER

for i, row in raw.iterrows():
    for j, col in enumerate(cols, 1):
        val = row[col]
        if col == "is_novelty_effect":
            val = 1 if val else 0
        c = ws.cell(i + 2, j, val)
        c.font = INPUT_FONT
        c.border = BORDER

ws.cell(11, 1, "Source: novelty_decay_pct_per_day and is_novelty_effect are computed via linear").font = SUBTITLE_FONT
ws.cell(12, 1, "regression on daily lift (scipy.stats.linregress) — see src/stats/trust_engine.py::novelty_effect_score.").font = SUBTITLE_FONT
ws.cell(13, 1, "Required N/arm uses a standard two-proportion power calculation (baseline=10%, MDE=2pp, alpha=0.05, power=80%).").font = SUBTITLE_FONT
for r in [11, 12, 13]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

for j, w in enumerate([14, 26, 14, 11, 12, 16, 17, 15, 17, 16, 16, 22, 15, 16, 12, 12], 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A2"

N = len(raw)
LAST_ROW = N + 1

# =====================================================================
# SHEET 2: Trust_Engine — every formula mirrors dax/measures.md exactly
# =====================================================================
ts = wb.create_sheet("Trust_Engine")
te_headers = [
    "Experiment ID", "Experiment Name", "Owner Team", "Control N", "Treatment N",
    "SRM Chi-Square", "SRM P-Value", "Is SRM?",
    "Control Conv Rate", "Treatment Conv Rate", "Absolute Lift", "Relative Lift %",
    "Pooled Rate", "Pooled SE", "Z-Stat", "P-Value", "Significant?",
    "Unpooled SE", "CI Lower", "CI Upper",
    "Control Guardrail Rate", "Treatment Guardrail Rate", "Guardrail P-Value", "Guardrail Breached?",
    "Active Users v1", "Active Users v2", "Definition Agreement %",
    "Required N/Arm", "Sample Size Adequate?",
    "Novelty Decay %/day", "Novelty Effect?",
    "SRM Pts (30)", "Definition Pts (25)", "Guardrail Pts (20)", "Sample Pts (15)", "Novelty Pts (10)",
    "TRUST SCORE", "VERDICT",
]
for j, h in enumerate(te_headers, 1):
    c = ts.cell(1, j, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER

for i in range(2, LAST_ROW + 1):
    r = i
    ri = f"Raw_Inputs!"
    # links to Raw_Inputs
    ts.cell(r, 1, f"={ri}A{r}").font = LINK_FONT
    ts.cell(r, 2, f"={ri}B{r}").font = LINK_FONT
    ts.cell(r, 3, f"={ri}C{r}").font = LINK_FONT
    ts.cell(r, 4, f"={ri}D{r}").font = LINK_FONT      # Control N
    ts.cell(r, 5, f"={ri}E{r}").font = LINK_FONT      # Treatment N

    D, E = f"D{r}", f"E{r}"
    # SRM chi-square: sum of ((observed-expected)^2/expected) for both arms, expected = total/2
    ts.cell(r, 6, f"=(({D}-(({D}+{E})/2))^2/(({D}+{E})/2))+(({E}-(({D}+{E})/2))^2/(({D}+{E})/2))").font = FORMULA_FONT
    ts.cell(r, 7, f"=CHIDIST(F{r},1)").font = FORMULA_FONT
    ts.cell(r, 8, f'=IF(G{r}<0.001,1,0)').font = FORMULA_FONT

    ts.cell(r, 9, f"={ri}F{r}/{D}").font = FORMULA_FONT   # control conv rate
    ts.cell(r, 10, f"={ri}G{r}/{E}").font = FORMULA_FONT  # treatment conv rate
    ts.cell(r, 11, f"=J{r}-I{r}").font = FORMULA_FONT     # absolute lift
    ts.cell(r, 12, f"=IFERROR(K{r}/I{r}*100,0)").font = FORMULA_FONT

    ts.cell(r, 13, f"=({ri}F{r}+{ri}G{r})/({D}+{E})").font = FORMULA_FONT  # pooled rate
    ts.cell(r, 14, f"=SQRT(M{r}*(1-M{r})*(1/{D}+1/{E}))").font = FORMULA_FONT  # pooled SE
    ts.cell(r, 15, f"=K{r}/N{r}").font = FORMULA_FONT  # z-stat
    ts.cell(r, 16, f"=2*(1-NORMSDIST(ABS(O{r})))").font = FORMULA_FONT  # p-value
    ts.cell(r, 17, f"=IF(P{r}<0.05,1,0)").font = FORMULA_FONT

    ts.cell(r, 18, f"=SQRT(I{r}*(1-I{r})/{D}+J{r}*(1-J{r})/{E})").font = FORMULA_FONT  # unpooled SE
    ts.cell(r, 19, f"=K{r}-1.96*R{r}").font = FORMULA_FONT  # CI lower
    ts.cell(r, 20, f"=K{r}+1.96*R{r}").font = FORMULA_FONT  # CI upper

    ts.cell(r, 21, f"={ri}H{r}/{D}").font = FORMULA_FONT  # control guardrail rate
    ts.cell(r, 22, f"={ri}I{r}/{E}").font = FORMULA_FONT  # treatment guardrail rate
    # guardrail two-proportion test (pooled)
    ts.cell(r, 23,
        f"=2*(1-NORMSDIST(ABS((V{r}-U{r})/SQRT((({ri}H{r}+{ri}I{r})/({D}+{E}))*(1-(({ri}H{r}+{ri}I{r})/({D}+{E})))*(1/{D}+1/{E})))))"
    ).font = FORMULA_FONT
    ts.cell(r, 24, f"=IF(AND(W{r}<0.05,V{r}>U{r}),1,0)").font = FORMULA_FONT

    ts.cell(r, 25, f"={ri}J{r}").font = LINK_FONT
    ts.cell(r, 26, f"={ri}K{r}").font = LINK_FONT
    ts.cell(r, 27, f"=MIN(Y{r},Z{r})/MAX(Y{r},Z{r})*100").font = FORMULA_FONT

    ts.cell(r, 28, f"={ri}L{r}").font = LINK_FONT
    ts.cell(r, 29, f"=IF(MIN({D},{E})>=AB{r},1,0)").font = FORMULA_FONT

    ts.cell(r, 30, f"={ri}M{r}").font = LINK_FONT
    ts.cell(r, 31, f"={ri}N{r}").font = LINK_FONT

    ts.cell(r, 32, f"=IF(H{r}=0,30,0)").font = FORMULA_FONT
    ts.cell(r, 33, f"=MIN(AA{r},100)/100*25").font = FORMULA_FONT
    ts.cell(r, 34, f"=IF(X{r}=0,20,0)").font = FORMULA_FONT
    ts.cell(r, 35, f"=IF(AC{r}=1,15,0)").font = FORMULA_FONT
    ts.cell(r, 36, f"=IF(AE{r}=0,10,0)").font = FORMULA_FONT
    ts.cell(r, 37, f"=SUM(AF{r}:AJ{r})").font = FORMULA_FONT
    ts.cell(r, 37).font = Font(name=FONT, bold=True, size=10)
    ts.cell(r, 38, f'=IF(AK{r}>=85,"TRUST",IF(AK{r}>=60,"TRUST WITH CAVEATS","DO NOT TRUST"))').font = Font(name=FONT, bold=True, size=10)

    for j in range(1, 39):
        ts.cell(r, j).border = BORDER

# number formats
pct_cols = [9, 10, 11, 12, 21, 22, 27]
for col in pct_cols:
    letter = get_column_letter(col)
    for r in range(2, LAST_ROW + 1):
        ts.cell(r, col).number_format = "0.00%" if col not in (12, 27) else "0.0"
for col in [7, 16, 23]:
    letter = get_column_letter(col)
    for r in range(2, LAST_ROW + 1):
        ts.cell(r, col).number_format = "0.000000"
for r in range(2, LAST_ROW + 1):
    ts.cell(r, 37).number_format = "0.0"

for j, w in enumerate([14, 26, 14] + [13]*35, 1):
    ts.column_dimensions[get_column_letter(j)].width = w
ts.freeze_panes = "D2"

# Conditional formatting on VERDICT column
verdict_range = f"AL2:AL{LAST_ROW}"
ts.conditional_formatting.add(verdict_range,
    CellIsRule(operator="equal", formula=['"TRUST"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
ts.conditional_formatting.add(verdict_range,
    CellIsRule(operator="equal", formula=['"TRUST WITH CAVEATS"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500", bold=True)))
ts.conditional_formatting.add(verdict_range,
    CellIsRule(operator="equal", formula=['"DO NOT TRUST"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))

score_range = f"AK2:AK{LAST_ROW}"
ts.conditional_formatting.add(score_range, ColorScaleRule(
    start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B"))

print("Raw_Inputs and Trust_Engine sheets built.")

# =====================================================================
# SHEET 3: Daily_Lift — wide format for easy multi-series line charting
# =====================================================================
dl = wb.create_sheet("Daily_Lift")
dl_cols = list(daily_wide.columns)  # day, EXP-001, EXP-002, ...
for j, h in enumerate(dl_cols, 1):
    c = dl.cell(1, j, h if h != "day" else "Day")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
for i, row in daily_wide.iterrows():
    for j, col in enumerate(dl_cols, 1):
        val = row[col]
        c = dl.cell(i + 2, j, None if pd.isna(val) else val)
        c.font = INPUT_FONT
        c.border = BORDER
        if col != "day":
            c.number_format = "0.00%"
dl_last_row = len(daily_wide) + 1
for j in range(1, len(dl_cols) + 1):
    dl.column_dimensions[get_column_letter(j)].width = 13
dl.freeze_panes = "B2"

# =====================================================================
# SHEET 4: Metric_Lineage
# =====================================================================
ml = wb.create_sheet("Metric_Lineage")
ml_cols = ["metric_name", "experiment_id", "definition_version", "definition_logic", "effective_start", "effective_end"]
ml_headers = ["Metric Name", "Experiment ID", "Version", "Definition Logic", "Effective Start", "Effective End"]
for j, h in enumerate(ml_headers, 1):
    c = ml.cell(1, j, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
for i, row in metric_defs.iterrows():
    for j, col in enumerate(ml_cols, 1):
        c = ml.cell(i + 2, j, row[col])
        c.font = INPUT_FONT
        c.border = BORDER
for j, w in enumerate([14, 14, 10, 55, 15, 15], 1):
    ml.column_dimensions[get_column_letter(j)].width = w
ml.cell(len(metric_defs) + 3, 1,
    "This sheet is documentation-as-a-dashboard: every experiment's active_user logic is traceable").font = SUBTITLE_FONT
ml.cell(len(metric_defs) + 4, 1,
    "to an exact definition and effective date range. Source: sql/views/analysis_views.sql :: v_user_event_counts").font = SUBTITLE_FONT

# =====================================================================
# SHEET 5: Dashboard — KPI cards + charts, all formula-driven
# =====================================================================
db = wb.create_sheet("Dashboard", 0)
db.sheet_view.showGridLines = False

db["B2"] = "Experiment Trust Engine"
db["B2"].font = TITLE_FONT
db["B3"] = "Can we actually trust this A/B test result? Live Excel model — every number below is a formula, not a hardcoded value."
db["B3"].font = SUBTITLE_FONT
db.merge_cells("B2:K2")
db.merge_cells("B3:K3")

kpi_specs = [
    ("TRUST", f'=COUNTIF(Trust_Engine!AL2:AL{LAST_ROW},"TRUST")', "2E7D32"),
    ("TRUST WITH CAVEATS", f'=COUNTIF(Trust_Engine!AL2:AL{LAST_ROW},"TRUST WITH CAVEATS")', "B7791F"),
    ("DO NOT TRUST", f'=COUNTIF(Trust_Engine!AL2:AL{LAST_ROW},"DO NOT TRUST")', "C0392B"),
    ("AVG TRUST SCORE", f'=AVERAGE(Trust_Engine!AK2:AK{LAST_ROW})', "1F2937"),
]
col_start = 2
for i, (label, formula, color) in enumerate(kpi_specs):
    c0 = col_start + i * 3
    cell_range_val = db.cell(5, c0)
    db.merge_cells(start_row=5, start_column=c0, end_row=7, end_column=c0 + 1)
    val_cell = db.cell(5, c0, formula)
    val_cell.font = Font(name=FONT, bold=True, size=28, color=color)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.number_format = "0.0" if "AVG" in label else "0"
    db.merge_cells(start_row=8, start_column=c0, end_row=8, end_column=c0 + 1)
    lbl_cell = db.cell(8, c0, label)
    lbl_cell.font = Font(name=FONT, size=9, color="6B7280", bold=True)
    lbl_cell.alignment = Alignment(horizontal="center")
    for rr in range(5, 9):
        for cc in range(c0, c0 + 2):
            db.cell(rr, cc).border = BORDER

# Table: experiment / trust score / verdict, sourced live from Trust_Engine
db["B11"] = "Experiment"
db["D11"] = "Team"
db["E11"] = "Trust Score"
db["F11"] = "Verdict"
db["G11"] = "Key Reason"
for col in ["B", "D", "E", "F", "G"]:
    c = db[f"{col}11"]
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
db.merge_cells("G11:K11")

for i in range(N):
    r = 12 + i
    src = 2 + i
    db.cell(r, 2, f"=Trust_Engine!B{src}").font = LINK_FONT
    db.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    db.cell(r, 4, f"=Trust_Engine!C{src}").font = LINK_FONT
    db.cell(r, 5, f"=Trust_Engine!AK{src}").font = Font(name=FONT, bold=True)
    db.cell(r, 5).number_format = "0.0"
    db.cell(r, 6, f"=Trust_Engine!AL{src}").font = Font(name=FONT, bold=True)
    db.cell(r, 7, f"=Raw_Inputs!A{src}").font = LINK_FONT  # placeholder, replaced below
    db.merge_cells(start_row=r, start_column=7, end_row=r, end_column=11)
    for cc in range(2, 12):
        db.cell(r, cc).border = BORDER

# Replace col G with a real reasons formula referencing SRM/guardrail/novelty flags
for i in range(N):
    r = 12 + i
    src = 2 + i
    db.cell(r, 7,
        f'=IF(Trust_Engine!H{src}=1,"SRM detected — allocation not trustworthy. ","")'
        f'&IF(Trust_Engine!X{src}=1,"Guardrail regressed. ","")'
        f'&IF(Trust_Engine!AE{src}=1,"Novelty decay detected. ","")'
        f'&IF(Trust_Engine!AC{src}=0,"Underpowered sample. ","")'
        f'&IF(AND(Trust_Engine!H{src}=0,Trust_Engine!X{src}=0,Trust_Engine!AE{src}=0,Trust_Engine!AC{src}=1),"No issues detected","")'
    ).font = Font(name=FONT, size=9, color="6B7280")

# Verdict conditional formatting on dashboard column F
db.conditional_formatting.add(f"F12:F{11+N}",
    CellIsRule(operator="equal", formula=['"TRUST"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
db.conditional_formatting.add(f"F12:F{11+N}",
    CellIsRule(operator="equal", formula=['"TRUST WITH CAVEATS"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500", bold=True)))
db.conditional_formatting.add(f"F12:F{11+N}",
    CellIsRule(operator="equal", formula=['"DO NOT TRUST"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))

# --- Bar chart: Trust Score by Experiment ---
bar = BarChart()
bar.type = "bar"
bar.title = "Trust Score by Experiment"
bar.y_axis.title = None
bar.x_axis.title = "Trust Score"
bar.height = 8
bar.width = 16
data = Reference(ts, min_col=37, min_row=1, max_row=LAST_ROW)
cats = Reference(ts, min_col=2, min_row=2, max_row=LAST_ROW)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.legend = None
db.add_chart(bar, "B23")

# --- Line chart: cumulative lift over time (novelty decay) ---
line = LineChart()
line.title = "Cumulative Treatment Lift Over Time (Novelty Check)"
line.y_axis.title = "Lift"
line.x_axis.title = "Day"
line.height = 8
line.width = 16
data2 = Reference(dl, min_col=2, max_col=len(dl_cols), min_row=1, max_row=dl_last_row)
cats2 = Reference(dl, min_col=1, min_row=2, max_row=dl_last_row)
line.add_data(data2, titles_from_data=True)
line.set_categories(cats2)
db.add_chart(line, "K23")

for j, w in enumerate([3, 16, 10, 12, 12, 12, 20, 10, 10, 10, 10], 1):
    db.column_dimensions[get_column_letter(j)].width = w
db.sheet_view.zoomScale = 90

# =====================================================================
# SHEET 0: README — first tab, explains the workbook
# =====================================================================
rm = wb.create_sheet("README", 0)
rm.sheet_view.showGridLines = False
rm["B2"] = "Experiment Trust Engine — Excel Edition"
rm["B2"].font = TITLE_FONT
lines = [
    "",
    "This workbook is a fully live, formula-driven rebuild of the Power BI report from the",
    "experiment-trust-engine GitHub repo — built in Excel because Power BI Desktop requires",
    "admin rights that weren't available. Every statistical calculation is a real Excel formula,",
    "not a pasted value, so it recalculates if you change the Raw_Inputs sheet.",
    "",
    "SHEETS",
    "  Dashboard      KPI cards + charts (Trust Score leaderboard, novelty-decay lift chart)",
    "  Trust_Engine   Every formula: SRM chi-square, two-proportion z-test, CI, composite Trust Score",
    "  Raw_Inputs     Hardcoded inputs (blue font) — raw counts pulled from the synthetic event log",
    "  Daily_Lift     Day-by-day lift per experiment, feeds the novelty-effect chart",
    "  Metric_Lineage Documentation of every 'active_user' metric definition version and its effective dates",
    "",
    "COLOR CONVENTION",
    "  Blue text   = hardcoded input (sourced from data/raw/*.csv via the Python generator)",
    "  Black text  = formula calculated within this workbook",
    "  Green text  = link to another sheet",
    "",
    "METHODOLOGY",
    "  Trust Score = 30 pts (no Sample Ratio Mismatch) + 25 pts (metric definition agreement)",
    "                + 20 pts (no guardrail regression) + 15 pts (adequate sample size)",
    "                + 10 pts (no novelty/decay effect). See docs/trust_score_methodology.md",
    "                in the repo for the full rationale.",
    "",
    "This workbook's formulas mirror dax/measures.md exactly (CHIDIST/NORMSDIST are the",
    "universally-supported equivalents of Power BI's CHISQ.DIST.RT/NORM.S.DIST), so migrating",
    "this model into Power BI later — once Desktop is installed — is a direct port.",
]
for i, line in enumerate(lines):
    c = rm.cell(4 + i, 2, line)
    c.font = Font(name=FONT, size=10, bold=line.isupper() and line.strip() != "", color="1F2937")
rm.column_dimensions["B"].width = 100

wb.save(Path(__file__).resolve().parent / "Experiment_Trust_Engine.xlsx")
print("Full workbook built: Experiment_Trust_Engine.xlsx")

